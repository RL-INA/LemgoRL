"""
Excel Parser script to read configurations from xls files.
@author: Arthur Müller (Fraunhofer IOSB-INA in Lemgo)
@email: arthur.mueller@iosb-ina.fraunhofer.de
"""

from openpyxl import load_workbook

def find_row_offset(ws, parameter_group):
    '''
    Finds the row-offset in excel sheet.

    :param ws: excel working sheet instance
    :param parameter_group: e.g. '[Algorithmic Parameters]'

    '''
    for r in range(1, ws.max_row):
        if ws.cell(row=r, column=2).value == parameter_group:
            row_offset = r + 1
    return row_offset


def find_col_offset(ws):
    '''
    Finds the column offset in excel sheet from where to infere the parameters.
    '''
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == 1:
            col_offset = c
    return col_offset - 1


def find_last_param_id(ws, column_offset):
    '''
    Finds the last id number in excel sheet
    '''
    for c in range(column_offset, ws.max_column + 1):
        if isinstance(ws.cell(1, c).value, int):
            param_id = ws.cell(1, c).value
    return param_id


def val_convert(val):
    '''
    Convert the type of params in excel sheet in corresponding python types.
    '''
    if isinstance(val, int) or isinstance(val, float):
        return val
    else:
        # bool convert
        if (val == 'False') or (val == 'false'):
            return False
        elif (val == 'True') or (val == 'true'):
            return True
        # None convert
        elif (val == 'None') or (val == 'none'):
            return None
        # list convert
        elif val[0] == '[':
            return [int(_) for _ in val[1:-1].split(',')]
        elif val == '{}':
            return {}
        else:
            # string values
            return val


def parse_algo_parameters(excel_config_path, param_group, desired_id='last'):
    '''
    Parses the parameters specified in excel_config_path with the id given in desired_id.
    Returns dictionaries for algorithm-training, sumo-Simulation and Markov Decision Process

    :param excel_config_path: path to config-xlsx-file with all
    :param desired_id: if 'last', then the function infers the last/highest id in the config-xlsx-file,
    else it must be int
    :return:
    '''
    parameter_group_mapping = {
        'algo': '[Algorithmic Parameters]',
        'model': '[Model Parameters]',
        'sumo': '[sumo configuration Parameters]',
        'mdp': '[MDP Parameters]',
    }
    wb = load_workbook(filename=excel_config_path)
    ws = wb["Tabelle1"]

    COLUMN_OFFSET = find_col_offset(ws)
    ROW_OFFSET = find_row_offset(ws, parameter_group=parameter_group_mapping[param_group])
    column = find_last_param_id(ws, COLUMN_OFFSET) if desired_id == 'last' else int(desired_id)

    algo_dict = {}
    for r in range(ROW_OFFSET, ws.max_row + 1):
        param = ws.cell(row=r, column=2).value
        if param is None:
            break
        algo_dict[ws.cell(row=r, column=2).value] = val_convert(ws.cell(row=r, column=COLUMN_OFFSET + column).value)

    # Dict-Correction for exploration params
    try:
        if param_group == 'algo':
            if algo_dict['type'] == 'EpsilonGreedy':
                exploration_dict = {
                    'exploration_config': {
                        'type': algo_dict['type'],
                        'initial_epsilon': algo_dict['initial_epsilon'],
                        'final_epsilon': algo_dict['final_epsilon'],
                        'epsilon_timesteps': algo_dict['epsilon_timesteps'],
                    }
                }
            else:
                exploration_dict = {
                    'exploration_config': {
                        'type': algo_dict['type'],
                        'temperature': algo_dict['temperature'],
                    }
                }
            del algo_dict['type'], algo_dict['initial_epsilon'], algo_dict['final_epsilon'], algo_dict['epsilon_timesteps'], \
                algo_dict['temperature']
            algo_dict = {**algo_dict, **exploration_dict}
    except:
        pass
    return algo_dict




# config = {}
# for r in range(ROW_OFFSET, ws.max_row + 1):
#     config[ws.cell(row=r, column=2).value] = ws.cell(row=r, column=COLUMN_OFFSET + column).value


# # Training
# STOP_CRITERION_EPOCHS = config['STOP_CRITERION_EPOCHS']
# LEARNING_RATE = float_conv(config['LEARNING_RATE'])
# USE_LR_SCHEDULER = bool_conv(config['USE_LR_SCHEDULER'])
# CLIPPING_VALUE = none_conv(config["CLIPPING_VALUE"])  # Gradient Clipping Value
#
# DEVICE = torch.device('cpu')  # 'cuda' oder 'cpu'
# MODEL_PATH_SAVE = fr"./trained_agents/policy_nn_parameters_{SIMULATION_ID}"
# MODEL_PATH_LOAD = fr"./trained_agents/policy_nn_parameters_{SIMULATION_ID}_2000epochs.dat"
#
# # NN-Parameter-Loading
# LOAD_PARAMETERS = False
# if LOAD_PARAMETERS:
#     EPOCH_NR_LOAD_OFFSET = int(
#         re.search(r"_[(0-9)]*epochs.dat", MODEL_PATH_LOAD, ).group().strip('_').strip('epochs.dat'))
# else:
#     EPOCH_NR_LOAD_OFFSET = 0
#
# # RL parameters
# GAMMA = float_conv(config["GAMMA"])
# EPSILON_START = float_conv(config["EPSILON_START"])
# EPSILON_FINAL = float_conv(config["EPSILON_FINAL"])
# EPSILON_ITERATIONS_TO_FINAL = config["EPSILON_ITERATIONS_TO_FINAL"]
#
# # DQN-Tweaks
# N_STEP = config["N_STEP"]  # N-step DQN
# DOUBLE_DQN = bool_conv(config["DOUBLE_DQN"])  # Double_DQN
#
# # traCI/sumo parameters
# NOGUI = True
# SUMOGUI_START_WITHOUT_AFFIRMATION = True
# FLAG_FIRST_TIME_TRACI_START = True
#
# TIME_TO_TELEPORT = config["TIME_TO_TELEPORT"]  # negative Werte deaktivieren den Vehicle-Teleport
# WAITING_TIME_MEMORY = config[
#     "WAITING_TIME_MEMORY"]  # Die letzten x timesteps, die für die Berechnung der AccumulatedWaitingTime eines vehicles verwendet werden
# MAX_TIMESTEPS_PER_SIMULATION_RUN = config["MAX_TIMESTEPS_PER_SIMULATION_RUN"]
#
# YELLOW_TIME = config["YELLOW_TIME"]
# DELTA_T_GREEN = config["DELTA_T_GREEN"]  # in seconds
# ALPHA = float_conv(
#     config["ALPHA"])  # wird mit wait multipliziert und dient als Balancing zwischen queue_length und wait
# ACCEPTABLE_TIME = config[
#     "ACCEPTABLE_TIME"]  # Time which is acceptable for a driver to wait before a red light in seconds.
# MAX_NUMBER_VEHICLES_PER_DETECTOR = config["MAX_NUMBER_VEHICLES_PER_DETECTOR"]
#
# # Standardizing queue_length
# QL_MEAN = none_conv(config['QL_MEAN'])
# QL_STD = none_conv(config['QL_STD'])
#
# # DQN-specific
# EXPONENT_LAYER1 = config["EXPONENT_LAYER1"]
# EXPONENT_LAYER2 = config["EXPONENT_LAYER2"]
# UPDATE_TARGET_NET = config["UPDATE_TARGET_NET"]  # Update target_net every x iterations
# SOFT_UPDATE = bool_conv(config["SOFT_UPDATE"])
# TAU_SOFT_UPDATE = config["TAU_SOFT_UPDATE"]
# REPLAY_BUFFER_SIZE = config[
#     "REPLAY_BUFFER_SIZE"]  # int(15 * MAX_TIMESTEPS_PER_SIMULATION_RUN / 5) # Je größer der Buffer, desto länger dauert eine Iteration. Ab einer gewissen Größe merkt man das deutlich.
# BATCH_SIZE = config["BATCH_SIZE"]
# MIN_BUFFER_SIZE_FOR_TRAINING = BATCH_SIZE + N_STEP - 1  # config['MIN_BUFFER_SIZE_FOR_TRAINING']
#
# if DOUBLE_DQN & SOFT_UPDATE:
#     raise ValueError("Double DQN und Soft Update dürfen aus algorithmischen Gründen nicht gleichzeitig aktiv sein.")
#
# ACTION_MAP = {0: 0, 1: 2}  # -1 =idle
# ACTION_RANDOM_SAMPLING = list(ACTION_MAP.keys())
#
# DETECTOR_LENGTH_PER_LANE = config['DETECTOR_LENGTH_PER_LANE']
# # ROADS = [
# #     'west_inc', 'east_inc', 'north_inc', 'south_inc'
# # ]
#
# # LANES = [
# #     "west_inc_0", "west_inc_1", "west_out_0", "west_out_1",
# #     "east_inc_0", "east_inc_1", "east_out_0", "east_out_1",
# #     "north_inc_0", "north_inc_1", "north_out_0", "north_out_1",
# #     "south_inc_0", "south_inc_1", "south_out_0", "south_out_1",
# # ]
#
# DETECTOR_IDs = [
#     "e2_west_inc_0", "e2_west_inc_1",
#     "e2_east_inc_0", "e2_east_inc_1",
#     "e2_north_inc_0", "e2_north_inc_1",
#     "e2_south_inc_0", "e2_south_inc_1",
# ]
#
#
# # CHANGE_PHASE = {
# #     'west_inc': 1,
# #     'east_inc': 1,
# #     'north_inc': 3,
# #     'south_inc': 3,
# # }
